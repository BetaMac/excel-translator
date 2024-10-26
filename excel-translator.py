from openpyxl import load_workbook
from googletrans import Translator
import time

def translate_languages(file_path, target_language='es'):
    """
    Translates text from column B (Primary Language) to column C (Translation) in an Excel file.
    
    Parameters:
    file_path (str): Path to the Excel file
    target_language (str): Target language code (e.g., 'es' for Spanish)
    """
    # Load the workbook and select the active sheet
    wb = load_workbook(file_path)
    sheet = wb.active
    
    # Initialize translator
    translator = Translator()
    
    # Get the maximum row count
    max_row = sheet.max_row
    
    print(f"Starting translation of {max_row} rows...")
    print("Reading from Column B (Primary Language) and writing to Column C (Translation)")
    
    # Iterate through rows, starting from row 2 to skip header if present
    for row in range(2, max_row + 1):
        # Get source text from column B
        primary_text = sheet[f"B{row}"].value
        
        # Skip empty cells
        if not primary_text:
            continue
            
        try:
            # Add small delay to avoid hitting API limits
            time.sleep(0.5)
            
            # Translate text
            translation = translator.translate(
                primary_text,
                dest=target_language
            )
            
            # Write translation to column C
            sheet[f"C{row}"] = translation.text
            
            # Print progress
            if row % 5 == 0:
                print(f"Processed {row} rows...")
                print(f"Latest translation: {primary_text} → {translation.text}")
                
        except Exception as e:
            print(f"Error translating row {row}: {str(e)}")
            continue
    
    # Save the workbook with a new filename to preserve original
    new_filename = file_path.replace('.xlsx', '_translated.xlsx')
    wb.save(new_filename)
    print(f"Translation completed! Saved as: {new_filename}")

# Updated to use your specific file
if __name__ == "__main__":
    translate_languages(
        file_path="spanish.xlsx",
        target_language='es'
    )